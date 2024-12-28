import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font

def run():
    # Set the title of the Streamlit application
    st.title("Amazon Inventory Ledger")

    # Allow the user to upload a CSV file
    uploaded_file = st.file_uploader("Upload your CSV file", type=["csv"])

    if uploaded_file:
        try:
            # Read the uploaded CSV file into a Pandas DataFrame
            data = pd.read_csv(uploaded_file)

            # Define the required columns that must be in the uploaded file
            required_columns = {"ASIN", "Event Type", "Quantity", "MSKU", "Title"}
            # Check if all required columns are present in the DataFrame
            if not required_columns.issubset(data.columns):
                st.error(f"Missing required columns: {required_columns - set(data.columns)}")
                return

            try:
                # Read the ASIN-to-SKU mapping file
                asin_to_sku_mapping = pd.read_csv("mappings/asin_to_sku_mapping.csv", sep="\t")
                # Verify that the mapping file contains the required columns
                if not {"ASIN", "Internal_SKU"}.issubset(asin_to_sku_mapping.columns):
                    st.error("Mapping file is missing required columns.")
                    return
            except FileNotFoundError:
                # Display an error message if the mapping file is not found
                st.error("Mapping file not found.")
                return

            # # Display the uploaded data in a table
            # st.subheader("Uploaded Data")
            # st.dataframe(data)

            # # Allow the user to filter data by ASIN
            # asin_filter = st.multiselect("Select ASIN(s) to filter", data["ASIN"].unique())
            # if asin_filter:
            #     data = data[data["ASIN"].isin(asin_filter)]

            # Section to display processed data
            st.subheader("Processed Data")

            # Create a pivot table for non-receipt event types, summing quantities
            non_receipt_pivot = data[data["Event Type"] != "Receipts"].pivot_table(
                index="ASIN", columns="Event Type", values="Quantity", aggfunc="sum", fill_value=0
            ).reset_index()

            # Add a column to sum all non-receipt event quantities
            non_receipt_pivot["Inventory_Out"] = non_receipt_pivot.drop(columns=["ASIN"]).sum(axis=1)

            # Calculate the total quantity for "Receipts" event type
            sum_receipts = data[data["Event Type"] == "Receipts"].groupby("ASIN")["Quantity"].sum().reset_index()
            sum_receipts.columns = ["ASIN", "Sum_Quantity_Receipts"]

            # Get a list of unique MSKUs for each ASIN
            msku_data = data.groupby("ASIN")["MSKU"].apply(
                lambda x: ", ".join(x.unique())).reset_index(name="Amazon_SKUs")

            # Get a list of unique Titles for each ASIN
            title_data = data.groupby("ASIN")["Title"].apply(
                lambda x: ", ".join(x.unique())).reset_index(name="Amazon_Titles")

            # Merge all processed data into a single DataFrame
            final_data = (
                non_receipt_pivot
                .merge(asin_to_sku_mapping, on="ASIN", how="left")
                .merge(sum_receipts, on="ASIN", how="left")
                .merge(msku_data, on="ASIN", how="left")
                .merge(title_data, on="ASIN", how="left")
                
            )
            columns = ["Internal_SKU"] + [col for col in final_data.columns if col != "Internal SKU"]
            final_data = final_data[columns]

            # # Identify ASINs missing Internal_SKU and prompt the user to add them
            # missing_skus = final_data[final_data["Internal_SKU"].isnull()]["ASIN"].tolist()
            # if missing_skus:
            #     st.warning("The following ASINs are missing Internal SKU mappings:")
            #     st.write(missing_skus)

            #     for asin in missing_skus:
            #         new_sku = st.text_input(f"Enter Internal SKU for ASIN {asin}:", key=asin)
            #         if new_sku:
            #             # Update the mapping DataFrame with the new SKU
            #             asin_to_sku_mapping = pd.concat(
            #                 [asin_to_sku_mapping, pd.DataFrame({"ASIN": [asin], "Internal_SKU": [new_sku]})],
            #                 ignore_index=True
            #             )

            # # Display the final processed data in a table
            # st.dataframe(final_data)

            # Function to generate an Excel file from the processed DataFrame
            def generate_excel(dataframe, column_widths=None, bold_columns=None, column_order=None):
                if column_order:
                    dataframe = dataframe[column_order]

                output = io.BytesIO()
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Inventory Ledger Sums"

                # Write header row with bold font
                for col_num, column_title in enumerate(dataframe.columns, 1):
                    sheet.cell(row=1, column=col_num, value=column_title).font = Font(bold=True)

                # Write data rows
                for row_num, row_data in enumerate(dataframe.itertuples(index=False), 2):
                    for col_num, value in enumerate(row_data, 1):
                        cell = sheet.cell(row=row_num, column=col_num, value=value)
                        # Make certain columns bold if specified
                        if bold_columns and col_num in bold_columns:
                            cell.font = Font(bold=True)

                # Adjust column widths if specified
                if column_widths:
                    for col_index, width in column_widths.items():
                        col_letter = sheet.cell(row=1, column=col_index).column_letter
                        sheet.column_dimensions[col_letter].width = width

                # Save the workbook to a BytesIO object
                workbook.save(output)
                output.seek(0)
                return output

            # Generate an Excel file for the processed data - Simple + add download button
            excel_data = generate_excel(
                final_data,
                column_order=["ASIN", "Internal_SKU", "Inventory_Out", "Amazon_Titles", "Amazon_SKUs"],
                column_widths={1: 16.5, 2: 16, 3: 13, 4: 150, 5: 150},
                bold_columns=[3]
            )
            st.download_button(
                label="Download Processed Data as Excel - Simple",
                data=excel_data,
                file_name="Inventory_ledger_simple.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            # Generate an Excel file for the processed data - Full + add download button
            excel_data = generate_excel(
                final_data,
                column_widths={1: 20, 2: 20, 3: 20, 4: 20, 5: 20, 6: 20, 7:20, 8:20, 9:29, 10:20, 11:20},
                bold_columns=[3]
            )
            st.download_button(
                label="Download Processed Data as Excel - Full",
                data=excel_data,
                file_name="Inventory_ledger_Full.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )



        except Exception as e:
            # Display an error message if an exception occurs
            st.error(f"An error occurred: {e}")
    else:
        # Prompt the user to upload a file if none is provided
        st.write("Please upload a CSV file to proceed.")

if __name__ == "__main__":
    run()
