import io
import datetime
from dateutil import tz
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

columns_to_sum = [
    "product sales",
    "product sales tax",
    "shipping credits",
    "shipping credits tax",
    "gift wrap credits",
    "giftwrap credits tax",
    "Regulatory Fee",
    "Tax On Regulatory Fee",
    "promotional rebates",
    "promotional rebates tax",
    "marketplace withheld tax",
    "selling fees",
    "fba fees",
    "other transaction fees",
    "other",
    "total",
]

# -------------------------------------------------------------------------
# Session State Initialization
# -------------------------------------------------------------------------
# We use Streamlit's session_state to persist data across interactions within the app
if "matching_transactions" not in st.session_state:
    st.session_state.matching_transactions = None
if "non_matching_transactions" not in st.session_state:
    st.session_state.non_matching_transactions = None

def clean_and_summarize_transactions(data, columns_to_sum=None, exclude_types=None,group_by=None):
    """
    Cleans and groups the data by 'type' and 'description' and calculates the sum for specified columns.
    Excludes transactions of specified types and handles numbers with commas.
    
    Parameters:
        data (pd.DataFrame): The transactions data.
        columns_to_sum (list): List of columns to sum.
        exclude_types (list): List of transaction types to exclude. Default is None.
    
    Returns:
        pd.DataFrame: A summarized DataFrame.
    """

    # Filter out specified transaction types, if any
    if exclude_types:
        data = data[~data['type'].isin(exclude_types)]
    # Fill NaN in the type column with a placeholder
    

    if columns_to_sum:
        # Remove commas and ensure specified columns are numeric
        for col in columns_to_sum:
            data[col] = data[col].replace({',': ''}, regex=True)  # Remove commas
            data[col] = pd.to_numeric(data[col], errors='coerce')  # Convert to numeric, coercing errors to NaN

        # Fill NaN in numeric columns with 0
        data[columns_to_sum] = data[columns_to_sum].fillna(0)

        # Group by 'type' and 'description' and sum specified columns
    
        summary = data.groupby(group_by)[columns_to_sum].sum().reset_index()
    else:
        summary = data
    return summary

def filter_transactions_by_month_year(dataframe, month, year, datetime_column='date/time'):
    """
    Cleans and filters transactions for a specific month and year, handling timezone issues.
    
    Parameters:
        dataframe (pd.DataFrame): The DataFrame containing the transactions.
        month (int): The month to filter by (1-12).
        year (int): The year to filter by (e.g., 2024).
        datetime_column (str): The name of the column containing date/time values.
        
    Returns:
        pd.DataFrame: Filtered DataFrame containing transactions from the specified month and year.
    """
    # Handle timezone abbreviations by removing them
    dataframe[datetime_column] = dataframe[datetime_column].str.replace(r" [A-Z]{3,4}$", "", regex=True)
    
    # Convert to datetime
    dataframe[datetime_column] = pd.to_datetime(dataframe[datetime_column], errors='coerce')
    
    # Filter the data for the specified month and year
    filtered_df = dataframe[
        (dataframe[datetime_column].dt.month == month) & 
        (dataframe[datetime_column].dt.year == year)
    ]
    
    return filtered_df


def validate_date_in_target(df: pd.DataFrame, date_column: str, report_name: str) -> pd.DataFrame:
    """
    Convert the specified date_column in the DataFrame to datetime.
    If errors occur, return an empty DataFrame and display an error.
    
    :param df: The input DataFrame.
    :param date_column: Column name that contains date information.
    :param report_name: Name of the report (for error messages).
    :return: The DataFrame with the date_column converted to datetime 
             or an empty DataFrame on error.
    """
    try:
        df[date_column] = pd.to_datetime(df[date_column], errors='coerce')
        return df
    except Exception as e:
        st.error(f"Error processing dates in {report_name}: {e}")
        return pd.DataFrame()


def combine_reports(reports: list) -> pd.DataFrame:
    """
    Combine multiple DataFrames into one using pandas.concat().
    
    :param reports: A list of pandas DataFrames to concatenate.
    :return: A single concatenated DataFrame.
    """
    return pd.concat(reports, ignore_index=True)


def read_text_or_csv(uploaded_file) -> pd.DataFrame:
    """
    Read a text or CSV file into a pandas DataFrame.
    Uses 'python' engine to automatically detect delimiters in CSV/TSV files.
    
    :param uploaded_file: The file uploaded by the user (Streamlit file-like object).
    :return: The resulting pandas DataFrame.
    """
    try:
        df = pd.read_csv(uploaded_file, sep=None, engine='python')
        return df
    except Exception as e:
        st.error(f"Error reading file: {e}")
        return pd.DataFrame()


def read_dynamic_header_report(uploaded_file, search_string: str = "date/time") -> pd.DataFrame:
    """
    Read a CSV (or text-based) report, dynamically detecting the header row by a given search string.
    
    :param uploaded_file: The file uploaded by the user (Streamlit file-like object).
    :param search_string: The string to search for in lines to identify the header row.
    :return: The resulting pandas DataFrame with correct headers.
    """
    try:
        # Decode the uploaded file content
        content = uploaded_file.read().decode('utf-8-sig')
        lines = content.splitlines()

        # Find the line index where the search_string occurs
        header_row = next(i for i, line in enumerate(lines) if search_string in line.lower())

        # Read the CSV from the detected header row onward
        df = pd.read_csv(io.StringIO(content), skiprows=header_row)
        return df
    except Exception as e:
        st.error(f"An error occurred while processing the uploaded file: {e}")
        return pd.DataFrame()


def run():
    """
    Main Streamlit application to process Amazon sales reports. 
    The user can upload several reports (Fulfillment, Unified Transaction, 
    Standard Orders Deferred Transaction, Invoiced Orders Deferred Transaction), 
    which are validated and combined. The combined data is then split into 
    matching and non-matching transactions (based on Fulfillment Report 
    order IDs) and summed for certain financial columns.
    """
    # ---------------------------------------------------------------------
    # Page Title and Instructions
    # ---------------------------------------------------------------------
    st.title("Amazon Sales Report Processor")
    st.write(
        "Upload your Amazon sales reports, specify the target month and year, "
        "and process them for accounting."
    )
    user_selected_date = st.date_input(
        "Select a date (we'll use only the month/year)"
    )
    
    # 2. Extract the month/year from user-selected date
    target_month = user_selected_date.month
    target_year = user_selected_date.year
    # ---------------------------------------------------------------------
    # File Uploaders
    # ---------------------------------------------------------------------
    excel_report = st.file_uploader("Upload Ecommerce Net of Plan (Optional)", type=['xlsx'])
    fulfillment_file = st.file_uploader("Upload Fulfillment Report:", type=["txt"])
    unified_transaction_file = st.file_uploader(
        "Upload Unified Transaction Report:", type=["csv", "xls", "xlsx"]
    )
    standard_orders_file = st.file_uploader(
        "Upload Standard Orders Deferred Transaction Report:", type=["csv", "xls", "xlsx"]
    )
    invoiced_orders_file = st.file_uploader(
        "Upload Invoiced Orders Deferred Transaction Report:", type=["csv", "xls", "xlsx"]
    )

    # ---------------------------------------------------------------------
    # Process the reports when the user clicks the "Process Reports" button
    # ---------------------------------------------------------------------
    if st.button("Process Reports"):
        if not all([fulfillment_file, unified_transaction_file, standard_orders_file, invoiced_orders_file]):
            st.error("Please upload all required reports.")
            return

        try:
            # -------------------------------------------------------------
            # Fulfillment Report
            # -------------------------------------------------------------
            st.subheader("Processing Fulfillment Report")
            fulfillment_df = read_text_or_csv(fulfillment_file)
            if fulfillment_df.empty:
                st.error("Fulfillment Report is empty or could not be processed.")
                return
            st.write("Fulfillment Report preview:", fulfillment_df.head())

            # Validate/convert date column for Fulfillment Report
            # Assumes the date is in the 3rd column (index = 2)
            try:
                date_col_name = fulfillment_df.columns[2]
                fulfillment_df = validate_date_in_target(
                    df=fulfillment_df, 
                    date_column=date_col_name, 
                    report_name="Fulfillment Report"
                )
            except Exception as e:
                st.error(f"Error processing Fulfillment Report date column: {e}")
                return

            # -------------------------------------------------------------
            # Unified Transaction Report
            # -------------------------------------------------------------
            st.subheader("Processing Unified Transaction Report")
            unified_transaction_df = read_dynamic_header_report(unified_transaction_file, search_string="date/time")
            # if not unified_transaction_df.empty:
            #     st.write(unified_transaction_df.head())

            # -------------------------------------------------------------
            # Standard Orders Deferred Transaction Report
            # -------------------------------------------------------------
            st.subheader("Processing Standard Orders Deferred Transaction Report")
            standard_orders_df = read_dynamic_header_report(standard_orders_file, search_string="date/time")
            # if not standard_orders_df.empty:
            #     st.write(standard_orders_df.head())

            # -------------------------------------------------------------
            # Invoiced Orders Deferred Transaction Report
            # -------------------------------------------------------------
            st.subheader("Processing Invoiced Orders Deferred Transaction Report")
            invoiced_orders_df = read_dynamic_header_report(invoiced_orders_file, search_string="date/time")
            # if not invoiced_orders_df.empty:
            #     st.write(invoiced_orders_df.head())

            # -------------------------------------------------------------
            # Combine the Unified, Standard Orders, and Invoiced Orders
            # into one transactions DataFrame
            # -------------------------------------------------------------
            combined_transactions = combine_reports([
                unified_transaction_df, 
                standard_orders_df, 
                invoiced_orders_df
            ])

            # st.write("Combined Transactions preview:", combined_transactions.head())
            
            # REMOVE COMMAS FROM DOLLAR VALUES
            # Find the index of the 'product sales' column
            start_col = combined_transactions.columns.get_loc('product sales')
            # Select all columns starting from 'product sales'
            columns_to_clean = combined_transactions.columns[start_col:]
            # Remove commas and convert to numeric for all selected columns
            for col in columns_to_clean:
                combined_transactions[col] = (
                    combined_transactions[col]
                    .astype(str)  # Ensure values are strings for str.replace
                    .str.replace(',', '', regex=True)  # Remove commas
                )
                combined_transactions[col] = pd.to_numeric(combined_transactions[col], errors='coerce')  # Convert to numeric
            # -------------------------------------------------------------
            # Provide a download link for the combined data
            # -------------------------------------------------------------
            csv_combined = combined_transactions.to_csv(index=False)
            st.download_button(
                label="Download Processed Data",
                data=csv_combined,
                file_name="processed_data.csv",
                mime="text/csv"
            )
            

            # -------------------------------------------------------------
            # Identify matching and non-matching transactions
            # -------------------------------------------------------------
            transaction_order_id_column = "order id"
            fulfillment_order_id_column = "amazon-order-id"
            type_column = "type"
            valid_types = ["Order"]  # Extend if needed, e.g., ["Order", "Liquidation"]

            # Validate columns exist
            required_combined_cols = [transaction_order_id_column, type_column]
            if all(col in combined_transactions.columns for col in required_combined_cols) and \
               fulfillment_order_id_column in fulfillment_df.columns:

                # Fulfillment order IDs
                fulfillment_order_ids = fulfillment_df[fulfillment_order_id_column]

                # Matching transactions based on 'order id' in Fulfillment and type
                matching_transactions = combined_transactions[
                    (combined_transactions[transaction_order_id_column].isin(fulfillment_order_ids)) &
                    (combined_transactions[type_column].isin(valid_types))
                ]

                # Non-matching transactions are everything else
                non_matching_transactions = combined_transactions[
                    ~(
                        (combined_transactions[transaction_order_id_column].isin(fulfillment_order_ids)) &
                        (combined_transactions[type_column].isin(valid_types))
                    )
                ]


                # ---------------------------------------------------------
                # Identify unexpected columns and sum known financial columns
                # ---------------------------------------------------------

                # Identify columns appearing after "product sales"
                if "product sales" in matching_transactions.columns:
                    columns_after_product_sales = matching_transactions.columns[
                        matching_transactions.columns.get_loc("product sales") + 1 :
                    ]
                    unexpected_columns = [
                        col for col in columns_after_product_sales if col not in columns_to_sum
                    ]

                    if unexpected_columns:
                        st.warning(
                            f"The following unexpected columns were found after 'product sales': "
                            f"{', '.join(unexpected_columns)}. These columns are not currently "
                            "accounted for in the calculations."
                        )
                else:
                    st.warning("Column 'product sales' is missing from the transactions.")
                # Convert columns_to_sum to numeric and sum them
                for col in columns_to_sum:
                    if col in matching_transactions.columns:
                        matching_transactions[col] = pd.to_numeric(matching_transactions[col], errors='coerce')
                    else:
                        st.warning(f"Column '{col}' is missing from the report.")

                
                summed_values = matching_transactions[columns_to_sum].sum(numeric_only=True)
                st.subheader("Summed Column Values")
                
                summed_df = summed_values.to_frame().reset_index()
                summed_df.columns = ['Charge Type', 'Amount']
                st.write("Matching Order Sales/Charges:", summed_df)
                
                unified_transaction_df['type'] = unified_transaction_df['type'].fillna('No Type')
                filtered_transactions_df = filter_transactions_by_month_year(unified_transaction_df,target_month,target_year)
                filtered_transactions_df = clean_and_summarize_transactions(filtered_transactions_df,exclude_types=['Transfer','Order'])
                filtered_transactions_df = combine_reports ([filtered_transactions_df, matching_transactions])
                df_checker = filtered_transactions_df
                # Make description uniform for all product transactions
                filtered_transactions_df.loc[filtered_transactions_df['type'].isin(['Order', 'Liquidations','Liquidations Adjustments','Refund']), 'description'] = '(items)'
                filtered_transactions_df['fulfillment'] = filtered_transactions_df['fulfillment'].fillna('none')
                filtered_transactions_df['description'] = filtered_transactions_df['description'].fillna('none')
                filtered_transactions_df = clean_and_summarize_transactions(filtered_transactions_df,columns_to_sum,group_by=['type', 'description','fulfillment'])

                df = filtered_transactions_df

                #load first sheet of excel file with P&L numbers
                fbm_sales = df.loc[(df['type'] == 'Order') & (df['fulfillment'] == 'Seller'), 'product sales'].sum()
                fba_sales = df.loc[(df['type'] == 'Order') & (df['fulfillment'] == 'Amazon'), 'product sales'].sum()

                # Avoid division by zero
                total_sales = fbm_sales + fba_sales
                fbm_percentage = (fbm_sales / total_sales) if total_sales > 0 else 0
                fba_percentage = (fba_sales / total_sales) if total_sales > 0 else 0

                fbm_returns = df.loc[(df['type'] == 'Refund') & (df['fulfillment'] == 'Seller'), 'product sales'].sum()
                fba_returns = df.loc[(df['type'] == 'Refund') & (df['fulfillment'] == 'Amazon'), 'product sales'].sum()

                fbm_commissions = df.loc[df['fulfillment'] == 'Seller', 'selling fees'].sum()
                fba_commissions = df.loc[df['fulfillment'] == 'Amazon', 'selling fees'].sum()

                advertising = df.loc[df['description'] == 'Cost of Advertising', 'total'].sum()

                fba_shipping = df['fba fees'].sum()

                fba_inbound_freight = next(
                    iter(df.loc[(df['type'] == 'FBA Inventory Fee') & 
                                (df['description'] == 'FBA Amazon-Partnered Carrier Shipment Fee'), 'total']), 0
                )

                service_fees_less_advertising = (
                    df.loc[df['type'] == 'Service Fee', 'total'].sum() - advertising
                )

                fbm_shipping_services = df.loc[df['type'] == 'Shipping Services', 'total'].sum()
                fba_adjustments = df.loc[df['type'] == 'Adjustment', 'total'].sum()

                fba_storage_fees = next(
                    iter(df.loc[(df['type'] == 'FBA Inventory Fee') & 
                                (df['description'] == 'FBA storage fee'), 'total']), 0
                )

                fba_inventory_fees_other = (
                    df.loc[df['type'] == 'FBA Inventory Fee', 'total'].sum() - fba_storage_fees - fba_inbound_freight
                )

                fba_liquidations = next(
                    iter(df.loc[(df['type'] == 'Liquidations') | 
                                (df['type'] == 'Liquidations Adjustments'), 'total']), 0
                )

                SAFE_T_reimbursement = next(iter(df.loc[df['type'] == 'SAFE-T reimbursement', 'total']), 0)

                fbm_promotional_rebate = df.loc[df['fulfillment'] == 'Seller', 'promotional rebates'].sum()
                fba_promotional_rebate = df.loc[df['fulfillment'] == 'Amazon', 'promotional rebates'].sum()

                total_unaccounted = df['total'].sum() - (fbm_sales+fba_sales+fbm_returns+fba_returns+fbm_commissions+fba_commissions+advertising+ \
                +fba_shipping+fba_inbound_freight+service_fees_less_advertising+fbm_shipping_services \
                +fba_adjustments+fba_storage_fees+fba_inventory_fees_other+fba_liquidations \
                +SAFE_T_reimbursement+fbm_promotional_rebate+fba_promotional_rebate)


                if excel_report is not None:
                    pnl_wb = load_workbook(excel_report)
                    pnl_wb_s1 = pnl_wb['Summary']
                    append_row_summary = 1
                    append_row = 9
                    append_column = 3
                    report_name = datetime.datetime(target_year, target_month, 1).strftime("Ecommerce Net of Plan - %B %Y")
                    other_expenses_list = ["Service Fees","Shipping Services","FBA Inventory Adjustment",\
                                           "FBA Storage Fees","FBA Liquidations","FBA Inventory - Other Fees",\
                                            "SAFE-T reimbursement","Promotional Rebates","Unaccounted","Total"]
                    iteration=0
                    for line in other_expenses_list:
                       pnl_wb_s1.cell(35+iteration,4,line) 
                       iteration=iteration+1
                    pnl_wb_s1.cell(33,4,"Amazon Other Income/Expenses")
                    pnl_wb_s1.cell(34,5,"FBM")
                    pnl_wb_s1.cell(34,6,"FBA")
                    pnl_wb_s1.cell(44,5,"=SUM(E35:E43)")
                    pnl_wb_s1.cell(44,6,"=SUM(F35:F43)")
                    pnl_wb_s1.cell(21,2,"=E44")
                    pnl_wb_s1.cell(21,3,"=F44")
                else:
                    pnl_wb = load_workbook('mappings/amazon-pnl-template.xlsx')
                    pnl_wb_s1 = pnl_wb.active
                    append_row_summary = 0
                    append_row = 0
                    append_column = 0
                    report_name = datetime.datetime(target_year, target_month, 1).strftime("Amazon Net of Plan - %B %Y (Adjusted)")
                    pnl_wb_s1.cell(2,1,report_name)

                col_fba = 3
                col_fbm = 2
                pnl_wb_s1.cell(append_row_summary+4,col_fbm,fbm_sales)
                pnl_wb_s1.cell(append_row_summary+4,col_fba,fba_sales)
                pnl_wb_s1.cell(append_row_summary+5,col_fbm,fbm_returns)
                pnl_wb_s1.cell(append_row_summary+5,col_fba,fba_returns)
                pnl_wb_s1.cell(append_row_summary+11,col_fbm,fbm_commissions)
                pnl_wb_s1.cell(append_row_summary+11,col_fba,fba_commissions)
                pnl_wb_s1.cell(append_row_summary+12,col_fbm,fbm_percentage*advertising)
                pnl_wb_s1.cell(append_row_summary+12,col_fba,fba_percentage*advertising)
                pnl_wb_s1.cell(append_row_summary+13,col_fba,fba_shipping)
                pnl_wb_s1.cell(append_row_summary+15,col_fba,fba_inbound_freight)
                pnl_wb_s1.cell(append_row+26,append_column+col_fbm,fbm_percentage*service_fees_less_advertising)
                pnl_wb_s1.cell(append_row+26,append_column+col_fba,fba_percentage*service_fees_less_advertising)
                pnl_wb_s1.cell(append_row+27,append_column+col_fbm,fbm_shipping_services)
                pnl_wb_s1.cell(append_row+28,append_column+col_fba,fba_adjustments)
                pnl_wb_s1.cell(append_row+29,append_column+col_fba,fba_storage_fees)
                pnl_wb_s1.cell(append_row+30,append_column+col_fba,fba_liquidations)
                pnl_wb_s1.cell(append_row+31,append_column+col_fba,fba_inventory_fees_other)
                pnl_wb_s1.cell(append_row+32,append_column+col_fbm,SAFE_T_reimbursement)
                pnl_wb_s1.cell(append_row+33,append_column+col_fbm,fbm_promotional_rebate)
                pnl_wb_s1.cell(append_row+33,append_column+col_fba,fba_promotional_rebate)
                pnl_wb_s1.cell(append_row+34,append_column+col_fbm,fbm_percentage*total_unaccounted)
                pnl_wb_s1.cell(append_row+34,append_column+col_fba,fba_percentage*total_unaccounted)

                    
                #Add amazon filtered transaction in new sheet
                pnl_wb_s2 = pnl_wb.create_sheet(title="Amazon Transaction Details")
                pnl_wb_s2.append(df.columns.tolist())
                for row in df.itertuples(index=False, name=None):
                    pnl_wb_s2.append(row)

                output2 = io.BytesIO()
                pnl_wb.save(output2)
                output2.seek(0)  # Reset the stream's position to the beginning

                # Streamlit UI
                st.title("Download Excel File Example")

                # Add a download button
                st.download_button(
                    label="Download Excel file",
                    data=output2,
                    file_name=report_name +".xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

                # Group by 'type' and 'description' and sum numeric columns
                # summary = clean_and_summarize_transactions(filtered_transactions_df,columns_to_sum,['Transfer','Order'])

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    filtered_transactions_df.to_excel(writer, index=False, sheet_name='Filtered')
                output.seek(0) 
                st.download_button(
                    "Download unified transactions that are not orders for the target month",
                    data=output,
                    file_name="filtered_unified_transactions.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                st.download_button(
                    "Check df",
                    data=df_checker.to_csv(index=False),
                    file_name="df_checker.csv",
                    mime="text/csv",
                )
            else:
                st.error("Required columns not found in one or both DataFrames.")

            st.success("Reports processed successfully!")

        except Exception as e:
            st.error(f"An error occurred during processing: {e}")

   

# Run the Streamlit app
if __name__ == "__main__":
    run()
